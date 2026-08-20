from datetime import timedelta

from django.contrib.auth import get_user_model
from django.db.models import Avg, Count, F, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import serializers, viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.clientes.models import Cliente
from apps.equipos.models import Equipo
from apps.evaluaciones.models import EvaluacionServicio
from apps.instalaciones.models import Instalacion
from apps.mantenimientos.models import Mantenimiento
from apps.materiales.models import Material
from apps.pagos.models import Pago
from apps.servicios.models import OrdenServicio

User = get_user_model()


class DashboardSerializer(serializers.Serializer):
    pass


def _puede_ver_reportes(user):
    return user and user.is_authenticated and user.role in (
        'administrador', 'supervisor', 'almacen'
    )


class DashboardViewSet(viewsets.ViewSet):
    """Estadísticas generales del sistema (Dashboard operativo)."""
    serializer_class = DashboardSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request):
        if not _puede_ver_reportes(request.user):
            return Response(
                {'detail': 'No tienes permisos para consultar el dashboard.'}, status=403
            )
        data = self._resumen()
        return Response(data)

    def _resumen(self):
        from apps.notificaciones.services import recordatorios_mantenimiento
        from apps.tienda.models import Orden

        hoy = timezone.localdate()
        proximo = hoy + timedelta(days=30)

        # Los recordatorios de mantenimiento se generan al abrir el dashboard.
        try:
            recordatorios_mantenimiento()
        except Exception:
            pass

        servicios_por_estado = dict(
            OrdenServicio.objects.values_list('estado')
            .annotate(total=Count('id'))
            .values_list('estado', 'total')
        )
        instalaciones_por_estado = dict(
            Instalacion.objects.values_list('estado')
            .annotate(total=Count('id'))
            .values_list('estado', 'total')
        )

        total_tecnicos = User.objects.filter(role='tecnico').count()
        ocupados_ids = set(
            Instalacion.objects.filter(estado__in=['asignada', 'en_proceso'])
            .exclude(tecnico__isnull=True).values_list('tecnico_id', flat=True)
        ) | set(
            OrdenServicio.objects.filter(estado__in=['asignada', 'en_proceso'])
            .exclude(tecnico__isnull=True).values_list('tecnico_id', flat=True)
        )
        tecnicos_ocupados = len(ocupados_ids)

        ventas = Orden.objects.filter(
            estado__in=['confirmado', 'preparando', 'enviado', 'entregado', 'completada']
        ).aggregate(total=Sum('total'))['total'] or 0

        # Series para los gráficos del dashboard (RF-13/RNF-14).
        servicios_por_mes = [
            {'mes': str(r['mes'])[:7], 'total': r['total']}
            for r in OrdenServicio.objects.annotate(mes=TruncMonth('created_at'))
            .values('mes').annotate(total=Count('id')).order_by('mes')
        ]
        instalaciones_por_mes = [
            {'mes': str(r['mes'])[:7], 'total': r['total']}
            for r in Instalacion.objects.filter(estado='finalizada')
            .annotate(mes=TruncMonth('fecha_instalacion'))
            .values('mes').annotate(total=Count('id')).order_by('mes')
        ]
        materiales_stock_bajo_list = [
            {
                'id': m.id,
                'nombre': m.nombre,
                'codigo': m.codigo,
                'cantidad_disponible': float(m.cantidad_disponible),
                'stock_minimo': float(m.stock_minimo),
            }
            for m in Material.objects.filter(cantidad_disponible__lte=F('stock_minimo'))
            .order_by(F('cantidad_disponible') - F('stock_minimo'))[:10]
        ]

        return {
            'total_clientes': Cliente.objects.count(),
            'total_tecnicos': total_tecnicos,
            'tecnicos_disponibles': max(0, total_tecnicos - tecnicos_ocupados),
            'tecnicos_ocupados': tecnicos_ocupados,
            'total_usuarios': User.objects.count(),
            'total_equipos': Equipo.objects.count(),
            'total_instalaciones': Instalacion.objects.count(),
            'instalaciones_realizadas': instalaciones_por_estado.get('finalizada', 0),
            'instalaciones_pendientes': (
                instalaciones_por_estado.get('pendiente', 0)
                + instalaciones_por_estado.get('asignada', 0)
            ),
            'servicios_pendientes': (
                servicios_por_estado.get('pendiente', 0)
                + servicios_por_estado.get('asignada', 0)
            ),
            'servicios_en_proceso': servicios_por_estado.get('en_proceso', 0),
            'servicios_completados': servicios_por_estado.get('finalizada', 0),
            'servicios_por_estado': servicios_por_estado,
            'instalaciones_por_estado': instalaciones_por_estado,
            'mantenimientos_proximos': Mantenimiento.objects.filter(
                proxima_fecha__range=(hoy, proximo),
                estado__in=['pendiente', 'en_proceso'],
            ).count(),
            'mantenimientos_vencidos': Mantenimiento.objects.filter(
                proxima_fecha__lt=hoy,
                estado__in=['pendiente', 'en_proceso'],
            ).count(),
            'materiales_stock_bajo': Material.objects.filter(
                cantidad_disponible__lte=F('stock_minimo')
            ).count(),
            'materiales_count': Material.objects.count(),
            'total_pagos': Pago.objects.filter(estado='pagado').aggregate(
                total=Sum('monto')
            )['total'] or 0,
            'pagos_pendientes': Pago.objects.filter(estado='pendiente').aggregate(
                total=Sum('monto')
            )['total'] or 0,
            'ventas': ventas,
            'servicios_por_mes': servicios_por_mes,
            'instalaciones_por_mes': instalaciones_por_mes,
            'materiales_stock_bajo_list': materiales_stock_bajo_list,
            'calificacion_promedio': EvaluacionServicio.objects.aggregate(
                promedio=Avg('calificacion')
            )['promedio'],
        }

    def reportes_list(self, request):
        if not _puede_ver_reportes(request.user):
            return Response(
                {'detail': 'No tienes permisos para consultar reportes.'}, status=403
            )
        return Response({
            'dashboard': '/api/dashboard/',
            'servicios_por_tecnico': '/api/dashboard/servicios-por-tecnico/',
            'instalaciones_por_mes': '/api/dashboard/instalaciones-por-mes/',
            'servicios_por_mes': '/api/dashboard/servicios-por-mes/',
            'materiales_stock_bajo': '/api/dashboard/materiales-stock-bajo/',
        })

    def servicios_por_tecnico(self, request):
        if not _puede_ver_reportes(request.user):
            return Response({'detail': 'Sin permisos.'}, status=403)
        data = (
            OrdenServicio.objects
            .filter(tecnico__isnull=False)
            .values('tecnico__id', 'tecnico__user__first_name', 'tecnico__user__last_name')
            .annotate(total=Count('id'), finalizadas=Count('id', filter=Q(estado='finalizada')))
            .order_by('-total')
        )
        resultado = [
            {
                'tecnico_id': r['tecnico__id'],
                'nombre': f"{r['tecnico__user__first_name']} {r['tecnico__user__last_name']}".strip(),
                'total_ordenes': r['total'],
                'finalizadas': r['finalizadas'],
            }
            for r in data
        ]
        return Response(resultado)

    def instalaciones_por_mes(self, request):
        if not _puede_ver_reportes(request.user):
            return Response({'detail': 'Sin permisos.'}, status=403)
        data = (
            Instalacion.objects
            .filter(estado='finalizada')
            .annotate(mes=TruncMonth('fecha_instalacion'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        return Response([
            {'mes': str(r['mes'])[:7], 'total': r['total']} for r in data
        ])

    def servicios_por_mes(self, request):
        if not _puede_ver_reportes(request.user):
            return Response({'detail': 'Sin permisos.'}, status=403)
        data = (
            OrdenServicio.objects
            .annotate(mes=TruncMonth('created_at'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        return Response([
            {'mes': str(r['mes'])[:7], 'total': r['total']} for r in data
        ])

    def materiales_stock_bajo(self, request):
        if not _puede_ver_reportes(request.user):
            return Response({'detail': 'Sin permisos.'}, status=403)
        qs = Material.objects.filter(cantidad_disponible__lte=F('stock_minimo')).order_by(
            F('cantidad_disponible') - F('stock_minimo')
        )
        return Response([
            {
                'id': m.id,
                'nombre': m.nombre,
                'codigo': m.codigo,
                'cantidad_disponible': float(m.cantidad_disponible),
                'stock_minimo': float(m.stock_minimo),
            }
            for m in qs
        ])

    def _datos_exportacion(self, tipo):
        """Devuelve (título, cabeceras, filas) según el tipo de reporte."""
        if tipo == 'instalaciones':
            qs = Instalacion.objects.select_related('cliente', 'tecnico').all()
            return 'Instalaciones', ['#', 'Cliente', 'Técnico', 'Fecha programada', 'Estado'], [
                [i.id, i.cliente.nombre_completo, i.tecnico_nombre,
                 timezone.localtime(i.fecha_programada).strftime('%d/%m/%Y %H:%M') if i.fecha_programada else '',
                 i.get_estado_display()]
                for i in qs[:500]
            ]
        if tipo == 'servicios':
            qs = OrdenServicio.objects.select_related('cliente', 'tecnico').all()
            return 'Servicios', ['Orden', 'Cliente', 'Técnico', 'Tipo', 'Fecha', 'Estado'], [
                [o.numero, o.cliente.nombre_completo, o.tecnico_nombre,
                 o.get_tipo_servicio_display(), o.fecha.strftime('%d/%m/%Y'),
                 o.get_estado_display()]
                for o in qs[:500]
            ]
        if tipo == 'materiales':
            qs = Material.objects.all()
            return 'Materiales', ['Código', 'Nombre', 'Disponible', 'Mínimo', 'Precio'], [
                [m.codigo, m.nombre, float(m.cantidad_disponible),
                 float(m.stock_minimo), float(m.precio)]
                for m in qs[:500]
            ]
        if tipo == 'pagos':
            qs = Pago.objects.select_related('cliente').all()
            return 'Pagos', ['Cliente', 'Fecha', 'Monto', 'Método', 'Estado'], [
                [p.cliente.nombre_completo, p.fecha.strftime('%d/%m/%Y'),
                 float(p.monto), p.get_metodo_display(), p.get_estado_display()]
                for p in qs[:500]
            ]
        return 'General', ['Indicador', 'Valor'], [
            ['Clientes', Cliente.objects.count()],
            ['Técnicos', User.objects.filter(role='tecnico').count()],
            ['Instalaciones realizadas', Instalacion.objects.filter(estado='finalizada').count()],
            ['Servicios completados', OrdenServicio.objects.filter(estado='finalizada').count()],
            ['Pagos recibidos', float(Pago.objects.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or 0)],
            ['Materiales con stock bajo', Material.objects.filter(cantidad_disponible__lte=F('stock_minimo')).count()],
        ]

    def exportar(self, request):
        """Exporta reportes en PDF o Excel (formato=pdf|xlsx, tipo=general|instalaciones|servicios|materiales|pagos)."""
        if not _puede_ver_reportes(request.user):
            return HttpResponse('Sin permisos.', status=403)
        tipo = str(request.query_params.get('tipo', 'general')).lower()
        formato = str(request.query_params.get('formato', 'pdf')).lower()
        tipo = str(request.query_params.get('tipo', 'general')).lower()
        titulo, headers, rows = self._datos_exportacion(tipo)
        hoy = timezone.localdate().strftime('%d/%m/%Y')

        if formato == 'xlsx':
            import io
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = titulo
            ws.append(headers)
            for fila in rows:
                ws.append(fila)
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            for i, ancho in enumerate([max(len(str(x)) for x in [headers[i]] + [f[i] for f in rows]) + 2
                                       for i in range(len(headers))], start=1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(ancho, 45)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename=reporte_{tipo}_{hoy}.xlsx'
            return response

        # PDF
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(letter),
            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30,
        )
        estilos = getSampleStyleSheet()
        elementos = [
            Paragraph(f'REFRIMASTE — Reporte: {titulo}', estilos['Title']),
            Spacer(1, 6),
            Paragraph(f'Generado el {hoy}', estilos['Normal']),
            Spacer(1, 12),
        ]
        data = [headers] + [[str(c) for c in fila] for fila in rows]
        tabla = Table(data, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0284C7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elementos.append(tabla)
        doc.build(elementos)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=reporte_{tipo}_{hoy}.pdf'
        return response
